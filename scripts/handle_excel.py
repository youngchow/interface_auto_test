# -*- coding: utf-8 -*-
import openpyxl
from scripts.handle_config import do_config
from scripts.contants import TEST_DATA_FILES_PATH_CASES

class HandleExcel:
    def __init__(self,filename,sheetname = None):
        self.filename = filename
        self.sheetname = sheetname

    def get_case(self):
        workbook = openpyxl.load_workbook(self.filename)
        if self.sheetname is None:
            worksheet = workbook.active
        else:
            worksheet = workbook[self.sheetname]
        worksheet_content = tuple(worksheet.iter_rows(min_row=1,max_row=6,values_only=True))
        worksheet_title = worksheet_content[0]

        dict_list =[]
        for rows in tuple(worksheet.iter_rows(min_row=2,values_only=True)):
            dicts= dict(zip(worksheet_title,rows))
            dict_list.append(dicts)
        return dict_list

    def write_case(self,row,actual,result):
        other_workbook = openpyxl.load_workbook(self.filename)
        if self.sheetname is None:
            other_worksheet = other_workbook.active
        else:
            other_worksheet = other_workbook[self.sheetname]

        if isinstance(row,int) and (2<=row <= other_worksheet.max_row):
            other_worksheet.cell(row,do_config.get_int('excel','actual_col'),actual)
            other_worksheet.cell(row,do_config.get_int('excel','result_col'),result)
            other_workbook.save(filename=self.filename)




if __name__ == '__main__':
     hand_excel = HandleExcel(TEST_DATA_FILES_PATH_CASES)
     hand_read = hand_excel.get_case()
     print(hand_read[1])
     hand_excel.write_case(2,'zhaoyun','zhuge')